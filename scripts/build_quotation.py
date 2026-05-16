#!/usr/bin/env python3
"""Generate POC_報價單_杰隆印刷.xlsx for the Jlong Printing POC.

Single-sheet quotation with:
  [A] Asgard 平台訂閱（靜態文字，不入加總）
  [B] 一次性 POC 導入服務 — 三階段，人天 × 單價 自動加總
  [C] Phase 2 進階模組（另行報價，不入加總）
  [D] 加總區 — 一次性導入小計 + 5% 營業稅 + 含稅總價
  [E] 備註 + 簽核欄

Run from anywhere:
  python3 scripts/build_quotation.py
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Brand colours (mirrors index.html / flow.html)
NAVY = "0F3460"
LIGHT_GOLD = "F7EFD3"
SINDRI_PURPLE = "5B21B6"
MIMIR_GREEN = "065F46"
GREY_TEXT = "64748B"
HEADER_ROW_FILL = "EEF2FF"

FONT_NAME = "微軟正黑體"
DAILY_RATE = 20000


def _section_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _phase_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _money_format(cell):
    cell.number_format = '"NT$" #,##0'


def _write_item_row(ws, row, name, desc):
    """Write one B-section item: name, desc, 人天 (blank), 單價 (預填), 小計 (公式)."""
    ws.cell(row=row, column=1, value=name).font = Font(name=FONT_NAME, size=10)
    ws.cell(row=row, column=2, value=desc).font = Font(name=FONT_NAME, size=10, color=GREY_TEXT)
    # C: 人天 留空
    ws.cell(row=row, column=3, value=None).alignment = Alignment(horizontal="center")
    # D: 單價 預填
    d = ws.cell(row=row, column=4, value=DAILY_RATE)
    _money_format(d)
    d.alignment = Alignment(horizontal="right")
    # E: 小計 公式
    e = ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
    _money_format(e)
    e.alignment = Alignment(horizontal="right")


def _phase_subtotal(ws, row, label, start_row, end_row):
    ws.cell(row=row, column=1, value=label).font = Font(name=FONT_NAME, size=10, bold=True)
    e = ws.cell(row=row, column=5, value=f"=SUM(E{start_row}:E{end_row})")
    _money_format(e)
    e.font = Font(name=FONT_NAME, size=10, bold=True)
    e.alignment = Alignment(horizontal="right")


def build_workbook(out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "POC 報價單"

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # Row 1 — Title (merged A1:E1)
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "Asgard AI × 杰隆印刷　POC 報價單"
    t.font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # Row 2 — Meta
    meta = [
        ("A", f"報價日期：{date.today().strftime('%Y-%m-%d')}"),
        ("B", "有效期：30 日（自報價日起）"),
        ("C", "案件編號：________"),
        ("E", "聯絡人：________"),
    ]
    for col, val in meta:
        c = ws[f"{col}2"]
        c.value = val
        c.font = Font(name=FONT_NAME, size=10, color=GREY_TEXT)

    # ── [A] Asgard 平台訂閱 ──
    _section_header(ws, 4, "[A] Asgard 平台訂閱費（持續性月費，與下方一次性費用分開）")

    ws["A5"] = "方案"
    ws["B5"] = "Basic-Lite（依正式上線使用量可升級至 Regular NT$ 50,000/月 或 Plus NT$ 100,000/月）"
    ws["A6"] = "月費"
    ws["B6"] = "NT$ 20,000 / 月"
    for r in (5, 6):
        ws[f"A{r}"].font = Font(name=FONT_NAME, size=10, bold=True)
        ws[f"B{r}"].font = Font(name=FONT_NAME, size=10)

    ws["A7"] = "付款方式（擇一）"
    ws["A7"].font = Font(name=FONT_NAME, size=10, bold=True)

    ws["A8"] = "  ☐ 信用卡月扣"
    ws["B8"] = "NT$ 20,000 / 月（每月自動扣款）"
    ws["A9"] = "  ☐ 年付開立發票"
    ws["B9"] = "NT$ 240,000 / 年（一次開立，預付一年）"
    for r in (8, 9):
        ws[f"A{r}"].font = Font(name=FONT_NAME, size=10)
        ws[f"B{r}"].font = Font(name=FONT_NAME, size=10, color=GREY_TEXT)

    # ── [B] 一次性 POC 導入服務 ──
    _section_header(ws, 11, "[B] 一次性 POC 導入服務（4-6 週，單價 NT$ 20,000 / 人天 = 8 小時）")

    # Column headers row 12
    headers = [("A", "項目"), ("B", "工作內容"), ("C", "人天"), ("D", "單價"), ("E", "小計")]
    for col, val in headers:
        c = ws[f"{col}12"]
        c.value = val
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=HEADER_ROW_FILL)
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[12].height = 22

    phase_a = [
        ("雲端環境規劃與建置", "GCP / AWS 帳號設定、VPC、IAM、Secret 管理"),
        ("DB Schema 設計與建置", "訂單草稿 / 客戶 / 對話紀錄 / 知識庫 schema"),
        ("AP Server 部署", "後端服務容器化、反向代理、TLS 設定"),
        ("CI/CD 與基本監控初始化", "GitHub Actions、Log/Error 監控初始化"),
    ]
    phase_b = [
        ("Odin 工作流設計", "接單主流程、條件分支、轉接邏輯（於 Odin Studio 內設計）"),
        ("對話 Prompt 與欄位偵測邏輯", "LLM Prompt 工程、欄位缺口偵測、白話追問生成"),
        ("Chatbot 前端 UI 開發", "嵌入式對話框、訊息流、模擬圖預覽元件"),
        ("Session 管理 + 訂單草稿 API", "Session 持久化、訂單 JSON schema、後端 API"),
    ]
    phase_c = [
        ("圖檔上傳解析", "LLM Vision 串接、設計稿色數與尺寸建議"),
        ("AI 模擬圖生成", "Image Gen API 串接、貼標情境圖輸出"),
        ("舊客識別查詢", "Email / 電話比對、歷史訂單帶入流程"),
        ("情緒偵測 + 真人轉接通知", "情緒分類、Email / LINE Webhook 通知"),
        ("UAT 支援 + 上線部署", "Bug 修正、上線切換、煙霧測試"),
        ("教育訓練與文件交付", "業務操作訓練（1 小時）、技術交接文件"),
    ]

    # Compute row positions from item list lengths so that adding items
    # to any phase automatically updates all downstream row references.
    PHASE_A_HEADER_ROW = 13
    phase_a_first    = PHASE_A_HEADER_ROW + 1                 # 14
    phase_a_last     = phase_a_first + len(phase_a) - 1       # 17
    phase_a_subtotal = phase_a_last + 1                        # 18

    PHASE_B_HEADER_ROW = phase_a_subtotal + 1                  # 19
    phase_b_first    = PHASE_B_HEADER_ROW + 1                  # 20
    phase_b_last     = phase_b_first + len(phase_b) - 1        # 23
    phase_b_subtotal = phase_b_last + 1                        # 24

    PHASE_C_HEADER_ROW = phase_b_subtotal + 1                  # 25
    phase_c_first    = PHASE_C_HEADER_ROW + 1                  # 26
    phase_c_last     = phase_c_first + len(phase_c) - 1        # 31
    phase_c_subtotal = phase_c_last + 1                        # 32

    # Phase A: header, items, subtotal
    _phase_header(ws, PHASE_A_HEADER_ROW, "──── W1-W2  Phase A：基礎架構 + DB Schema + AP Server 部署 ────")
    for i, (name, desc) in enumerate(phase_a):
        _write_item_row(ws, phase_a_first + i, name, desc)
    _phase_subtotal(ws, phase_a_subtotal, "Phase A 階段小計", phase_a_first, phase_a_last)

    # Phase B: header, items, subtotal
    _phase_header(ws, PHASE_B_HEADER_ROW, "──── W3-W4  Phase B：核心對話流 + Odin 工作流建置 ────")
    for i, (name, desc) in enumerate(phase_b):
        _write_item_row(ws, phase_b_first + i, name, desc)
    _phase_subtotal(ws, phase_b_subtotal, "Phase B 階段小計", phase_b_first, phase_b_last)

    # Phase C: header, items, subtotal
    _phase_header(ws, PHASE_C_HEADER_ROW, "──── W5-W6  Phase C：進階功能 + UAT + 教育訓練 ────")
    for i, (name, desc) in enumerate(phase_c):
        _write_item_row(ws, phase_c_first + i, name, desc)
    _phase_subtotal(ws, phase_c_subtotal, "Phase C 階段小計", phase_c_first, phase_c_last)

    # ── [C] Phase 2 ──
    _section_header(ws, 34, "[C] Phase 2 進階模組（另行報價，不計入本次總價）")

    ws["A35"] = "Sindri Agent Hub"
    ws["B35"] = "多 Agent 編排（取代 Phase 1 由 Odin + 客製化程式組成的單一 LLM 對話）"
    ws["E35"] = "另行報價"
    ws["A35"].font = Font(name=FONT_NAME, size=10, bold=True, color=SINDRI_PURPLE)
    ws["B35"].font = Font(name=FONT_NAME, size=10, color=GREY_TEXT)
    ws["E35"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY_TEXT)
    ws["E35"].alignment = Alignment(horizontal="right")

    ws["A36"] = "Mimir Data Insight"
    ws["B36"] = "數據洞察儀表板（取代 Phase 1 的基本 DB 報表）"
    ws["E36"] = "另行報價"
    ws["A36"].font = Font(name=FONT_NAME, size=10, bold=True, color=MIMIR_GREEN)
    ws["B36"].font = Font(name=FONT_NAME, size=10, color=GREY_TEXT)
    ws["E36"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY_TEXT)
    ws["E36"].alignment = Alignment(horizontal="right")

    # ── [D] 加總區 ──
    _section_header(ws, 38, "[D] 一次性 POC 導入服務 加總")

    ws["A39"] = "一次性導入服務小計（未稅）"
    ws["E39"] = f"=E{phase_a_subtotal}+E{phase_b_subtotal}+E{phase_c_subtotal}"
    ws["A39"].font = Font(name=FONT_NAME, size=11, bold=True)
    ws["E39"].font = Font(name=FONT_NAME, size=11, bold=True)
    _money_format(ws["E39"])
    ws["E39"].alignment = Alignment(horizontal="right")

    ws["A40"] = "營業稅 5%"
    ws["E40"] = "=ROUND(E39*0.05,0)"
    ws["A40"].font = Font(name=FONT_NAME, size=11, bold=True)
    ws["E40"].font = Font(name=FONT_NAME, size=11, bold=True)
    _money_format(ws["E40"])
    ws["E40"].alignment = Alignment(horizontal="right")

    ws["A41"] = "含稅總價"
    ws["E41"] = "=E39+E40"
    ws["A41"].font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
    ws["A41"].fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    ws["E41"].font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
    ws["E41"].fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    _money_format(ws["E41"])
    ws["E41"].alignment = Alignment(horizontal="right")
    ws.row_dimensions[41].height = 28

    ws.merge_cells("A42:E42")
    ws["A42"] = "★ 本份總價僅為 POC 一次性導入服務費用；Asgard 平台訂閱費另計（見 [A] 區）"
    ws["A42"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY_TEXT)

    # ── [E] 備註 ──
    _section_header(ws, 44, "[E] 備註")

    notes = [
        "假設與排除：",
        "  1. 不含杰隆印刷既有 ERP / 生產系統的整合與對接",
        "  2. 不含 LINE 官方帳號 / FB Messenger 等通道串接（Phase 2 規劃）",
        "  3. 不含自動報價計算邏輯（需杰隆提供定價模型）",
        "  4. 雲端基礎設施費用（GCP / AWS）由客戶以實際使用量支付",
        "",
        "付款條件建議：簽約 30% / 中期驗收 40% / 最終驗收 30%",
        "報價有效期：自報價日起 30 日",
    ]
    for i, text in enumerate(notes):
        c = ws.cell(row=45 + i, column=1, value=text)
        c.font = Font(name=FONT_NAME, size=10, color=("000000" if text and not text.startswith(" ") else GREY_TEXT))

    # Signature
    ws["A54"] = "客戶簽章：______________________"
    ws["C54"] = "日期：______________"
    ws["A55"] = "Asgard AI：______________________"
    ws["C55"] = "日期：______________"
    for r in (54, 55):
        ws[f"A{r}"].font = Font(name=FONT_NAME, size=10)
        ws[f"C{r}"].font = Font(name=FONT_NAME, size=10)

    # Freeze panes — keep title row visible
    ws.freeze_panes = "A3"

    wb.save(out_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    out = repo_root / "POC_報價單_杰隆印刷.xlsx"
    build_workbook(out)
    print(f"Generated: {out}")


if __name__ == "__main__":
    main()
