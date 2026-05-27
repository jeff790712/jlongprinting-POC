"""
產生 杰隆印刷_AI接單系統_FunctionMap_Questions_v0.1.xlsx
比照 references/fsm/台新餐飲設備_派工系統_FunctionMap_Gantt_Questions_v0.4-2.xlsx 格式
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── 顏色 ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1E3A8A")   # 深藍 (Asgard Navy)
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
GRP_FILL   = PatternFill("solid", fgColor="DBEAFE")   # 淡藍 (群組行)
GRP_FONT   = Font(bold=True, color="1E3A8A", name="Calibri", size=10)
EVEN_FILL  = PatternFill("solid", fgColor="F0F7FF")
ODD_FILL   = PatternFill("solid", fgColor="FFFFFF")
STATUS_OK  = PatternFill("solid", fgColor="D1FAE5")   # 綠 ✅ POC
STATUS_P2  = PatternFill("solid", fgColor="E5E7EB")   # 灰 Phase 2
STATUS_PND = PatternFill("solid", fgColor="FEF9C3")   # 黃 待確認
P0_FILL    = PatternFill("solid", fgColor="FEE2E2")   # 紅 P0
P1_FILL    = PatternFill("solid", fgColor="FEF3C7")   # 橙 P1
P2_FILL    = PatternFill("solid", fgColor="D1FAE5")   # 綠 P2
BODY_FONT  = Font(name="Calibri", size=10)
BODY_FONT_B= Font(bold=True, name="Calibri", size=10)
WRAP       = Alignment(wrap_text=True, vertical="top")
CENTER     = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN       = Side(style="thin", color="D1D5DB")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_row(ws, cols: list[tuple], row=1):
    """寫入標題列並套用樣式"""
    for col_idx, (title, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=title)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def body_cell(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill or ODD_FILL
    c.font = font or BODY_FONT
    c.alignment = align or WRAP
    c.border = BORDER
    return c


def freeze_and_filter(ws, cell="A2"):
    ws.freeze_panes = cell
    ws.auto_filter.ref = ws.dimensions


# ══════════════════════════════════════════════════════════════════
# Sheet 1 — 功能地圖
# ══════════════════════════════════════════════════════════════════
FUNC_MAP = [
    # (id, 平台, 功能, epic, 子功能, 說明, 可使用角色, 狀態, 備註)
    # ── 網頁 Chatbot
    ("1.1",  "網頁 Chatbot（客戶端）", "1. 對話入口",    "Channel",       "網頁嵌入框",           "可嵌入杰隆官網的對話視窗，RWD 支援桌機與手機",                                                               "客戶",            "✅ POC",  "建議放在官網右下角浮動按鈕"),
    ("1.2",  "網頁 Chatbot（客戶端）", "1. 對話入口",    "Channel",       "意圖識別",             "識別來意：詢價／訂單追蹤／舊客補印／閒聊／轉接，導入對應對話流",                                               "客戶",            "✅ POC",  "LLM Router 節點"),
    ("1.3",  "網頁 Chatbot（客戶端）", "2. 詢價流程",    "Order Intake",  "開場引導",             "白話問候，建立 Session，判斷新／舊客身份",                                                                    "客戶",            "✅ POC",  ""),
    ("1.4",  "網頁 Chatbot（客戶端）", "2. 詢價流程",    "Order Intake",  "A 類欄位缺口偵測",     "即時追蹤 14 項必問欄位收集進度，依優先序追問尚未填寫的欄位",                                                    "客戶",            "✅ POC",  "核心引擎，Odin LLM + 客製化程式"),
    ("1.5",  "網頁 Chatbot（客戶端）", "2. 詢價流程",    "Order Intake",  "材質智能引導",         "白話描述（如「有光澤感」「銀色金屬感」）自動對應官網 8 種材質，不推薦目錄外材料",                               "客戶",            "✅ POC",  "材質清單見「材質對照表」分頁"),
    ("1.6",  "網頁 Chatbot（客戶端）", "2. 詢價流程",    "Order Intake",  "B 類條件欄位觸發",     "特定回答（如成捲、提到燙金）自動追問對應條件欄位（捲向、燙金規格等）",                                          "客戶",            "✅ POC",  ""),
    ("1.7",  "網頁 Chatbot（客戶端）", "2. 詢價流程",    "Order Intake",  "訂單草稿確認",         "A 類欄位收集完畢後，彙整條列摘要供客戶最終確認",                                                              "客戶",            "✅ POC",  ""),
    ("1.8",  "網頁 Chatbot（客戶端）", "2. 詢價流程",    "Order Intake",  "欄位收集進度顯示",     "對話側欄即時顯示已填／待填欄位狀態（進度追蹤器）",                                                            "客戶",            "✅ POC",  ""),
    ("1.9",  "網頁 Chatbot（客戶端）", "3. 圖檔上傳",    "File & AI",     "設計稿上傳",           "支援客戶上傳設計檔（JPG / PNG / PDF），觸發 AI 圖像分析",                                                     "客戶",            "✅ POC",  ""),
    ("1.10", "網頁 Chatbot（客戶端）", "3. 圖檔上傳",    "File & AI",     "AI 圖像分析",          "LLM Vision 解析設計稿，自動判斷色數、有無燙金/打凸工藝，回報客戶確認",                                          "客戶",            "✅ POC",  "Odin LLM Vision 節點"),
    ("1.11", "網頁 Chatbot（客戶端）", "4. AI 模擬圖",   "File & AI",     "貼標情境圖生成",       "依欄位（尺寸、材質、容器類型）生成貼標模擬情境圖；A 類欄位收集 50%+ 時主動詢問",                                "客戶",            "✅ POC",  "Odin Image Gen 節點；低解析度預覽含浮水印"),
    ("1.12", "網頁 Chatbot（客戶端）", "4. AI 模擬圖",   "File & AI",     "對話內模擬圖預覽",     "在對話框內顯示低解析度預覽圖（含浮水印），說明非正式打樣",                                                      "客戶",            "✅ POC",  ""),
    ("1.13", "網頁 Chatbot（客戶端）", "5. 舊客識別",    "CRM",           "Email / 電話比對",     "客戶輸入聯繫方式後，自動比對客戶資料庫，判斷是否為舊客",                                                        "客戶",            "✅ POC",  ""),
    ("1.14", "網頁 Chatbot（客戶端）", "5. 舊客識別",    "CRM",           "歷史訂單帶入",         "一年內舊客帶入上次訂單欄位，僅確認差異；超過一年則重新確認全部欄位",                                            "客戶",            "✅ POC",  ""),
    ("1.15", "網頁 Chatbot（客戶端）", "6. 補印流程",    "CRM",           "補印舊款式",           "舊客選擇補印同款，帶入舊資料，確認數量與出貨日期即可完成（≤ 3 輪對話）",                                        "客戶",            "✅ POC",  ""),
    ("1.16", "網頁 Chatbot（客戶端）", "7. 情緒偵測",    "Escalation",    "負面情緒判斷",         "連續 2 次偵測到明確不滿語句時，切換至轉接模式",                                                                "客戶",            "✅ POC",  "觸發閾值需與客戶確認"),
    ("1.17", "網頁 Chatbot（客戶端）", "7. 情緒偵測",    "Escalation",    "真人轉接通知",         "停止推進欄位收集，保留完整對話紀錄，發送通知給業務人員",                                                        "客戶 / 業務",     "✅ POC",  ""),
    ("1.18", "網頁 Chatbot（客戶端）", "7. 情緒偵測",    "Escalation",    "超規格轉接",           "材質／尺寸超範圍，或客戶明確要求真人時自動轉接",                                                                "客戶 / 業務",     "✅ POC",  ""),
    ("1.19", "網頁 Chatbot（客戶端）", "8. 多語言",      "Language",      "英文詢價",             "偵測英文輸入，以英文回應並將欄位對應至中文訂單格式",                                                            "客戶",            "✅ POC",  ""),
    ("1.20", "網頁 Chatbot（客戶端）", "9. LINE 頻道",   "Channel",       "LINE 詢價入口",        "透過 LINE 官方帳號進入 Odin 工作流",                                                                          "客戶",            "Phase 2", "需開通 LINE Business 帳號"),
    # ── Odin 工作流
    ("2.1",  "Odin 工作流（後端）",   "10. 工作流引擎",  "Workflow",      "Session 管理",         "每筆對話建立獨立 Session，跨輪追蹤對話狀態與欄位填寫進度",                                                      "系統",            "✅ POC",  ""),
    ("2.2",  "Odin 工作流（後端）",   "10. 工作流引擎",  "Workflow",      "對話節點編排",         "Odin 零程式碼設計完整詢價路徑、條件分支邏輯與轉接觸發規則",                                                      "系統",            "✅ POC",  ""),
    ("2.3",  "Odin 工作流（後端）",   "10. 工作流引擎",  "Workflow",      "圖檔解析串接",         "對接 LLM Vision API 解析客戶上傳設計稿",                                                                       "系統",            "✅ POC",  ""),
    ("2.4",  "Odin 工作流（後端）",   "10. 工作流引擎",  "Workflow",      "模擬圖生成串接",       "對接 Image Gen API，依訂單欄位生成貼標情境圖",                                                                  "系統",            "✅ POC",  ""),
    ("2.5",  "Odin 工作流（後端）",   "11. 通知機制",    "Notification",  "Email 通知",           "訂單草稿完成或觸發人工轉接時，自動發送 Email 通知業務人員",                                                      "業務",            "✅ POC",  "Odin HTTP 節點"),
    ("2.6",  "Odin 工作流（後端）",   "11. 通知機制",    "Notification",  "LINE 群組通知",        "推送通知至業務 LINE 群組",                                                                                     "業務",            "Phase 2", ""),
    # ── 資料層
    ("3.1",  "資料層（AP Server / DB）", "12. 訂單管理", "Data",          "訂單草稿儲存",         "結構化 JSON 儲存對話完成後的訂單草稿，欄位格式對應現有 Excel 訂單追蹤表",                                         "系統 / 業務",     "✅ POC",  "欄位定義見「欄位定義」分頁"),
    ("3.2",  "資料層（AP Server / DB）", "12. 訂單管理", "Data",          "訂單草稿匯出",         "訂單草稿可匯出為 JSON / CSV，供業務匯入現有 Excel",                                                             "業務",            "✅ POC",  ""),
    ("3.3",  "資料層（AP Server / DB）", "13. 客戶資料", "Data",          "客戶資料庫",           "Email / 電話索引儲存客戶基本資料與歷史訂單，支援舊客識別查詢",                                                    "系統",            "✅ POC",  "個資需加密儲存"),
    ("3.4",  "資料層（AP Server / DB）", "14. 對話紀錄", "Data",          "完整 Log 保存",        "儲存每筆對話完整紀錄（含時間戳），供品質分析與後續訓練使用",                                                      "系統 / 管理員",   "✅ POC",  "保留期限需與客戶確認"),
    ("3.5",  "資料層（AP Server / DB）", "15. 系統管理", "Infra",         "AP Server / 容器化",   "後端 API 服務、Session 管理、容器化部署與基本運行監控",                                                          "系統管理員",      "✅ POC",  ""),
    ("3.6",  "資料層（AP Server / DB）", "16. 數據分析", "Analytics",     "訂單轉換儀表板",       "詢價完成率、材質偏好趨勢、客群分析儀表板",                                                                       "管理員 / 業務主管", "Phase 2", "Phase 2 Mimir Data Insight"),
    # ── 管理後台
    ("4.1",  "管理後台（業務端）",    "17. 訂單確認",    "Back Office",   "草稿列表",             "查看所有完成詢價的訂單草稿清單，含欄位填寫完整度指標",                                                            "業務 / 業務主管", "✅ POC",  ""),
    ("4.2",  "管理後台（業務端）",    "17. 訂單確認",    "Back Office",   "確認 / 備註 / 退回",   "業務人員確認草稿、補充備註或退回，確認後觸發通知客戶",                                                            "業務",            "✅ POC",  ""),
    ("4.3",  "管理後台（業務端）",    "18. 轉接管理",    "Back Office",   "人工接管紀錄",         "顯示觸發情緒偵測或超規格轉接的對話清單，供業務後續跟進",                                                          "業務 / 業務主管", "✅ POC",  ""),
    ("4.4",  "管理後台（業務端）",    "19. 使用者管理",  "Back Office",   "帳號與權限管理",       "業務人員帳號建立、角色設定、停用",                                                                                "系統管理員",      "待確認",  "是否列入 POC 需雙方對齊"),
    ("4.5",  "管理後台（業務端）",    "20. 報表",        "Back Office",   "訂單統計報表",         "月/季訂單量、熱門材質、完成率等統計報表",                                                                         "業務主管",        "Phase 2", ""),
    # ── 報價引擎
    ("5.1",  "報價引擎（獨立模組）",  "21. 自動報價",    "Pricing",       "即時報價計算",         "欄位收集完成後自動輸出報價金額，需杰隆提供完整定價邏輯",                                                          "客戶 / 業務",     "Phase 2", "需杰隆提供定價邏輯文件"),
    ("5.2",  "報價引擎（獨立模組）",  "21. 自動報價",    "Pricing",       "高金額轉人工",         "超過設定門檻金額時，自動轉業務議價",                                                                             "業務",            "Phase 2", "門檻金額需與客戶確認"),
]


# ══════════════════════════════════════════════════════════════════
# Sheet 2 — 對話流程清單
# ══════════════════════════════════════════════════════════════════
FLOW_LIST = [
    # (編號, 流程名稱, 觸發條件, 使用角色, 對話步驟上限, 狀態, 備註)
    ("FL001", "新客詢價完整流程",     "新客 + 意圖=詢價",            "客戶",  "≤ 15 輪",  "✅ POC",  ""),
    ("FL002", "舊客補印同款",         "舊客識別成功 + 意圖=補印",    "客戶",  "≤ 3 輪",   "✅ POC",  "帶入歷史訂單欄位"),
    ("FL003", "舊客修改規格重新詢價", "舊客識別成功 + 意圖=詢價",    "客戶",  "≤ 10 輪",  "✅ POC",  "帶入欄位但允許修改"),
    ("FL004", "圖檔上傳 + AI 分析",   "詢價流程中，客戶上傳設計稿",  "客戶",  "—",        "✅ POC",  "觸發 LLM Vision 節點"),
    ("FL005", "AI 貼標模擬圖生成",    "A 類欄位收集 50%+",           "客戶",  "—",        "✅ POC",  "主動詢問是否生成"),
    ("FL006", "情緒偵測觸發轉接",     "連續 2 次負面情緒語句",       "客戶",  "—",        "✅ POC",  "發送 Email 通知業務"),
    ("FL007", "超規格轉接",           "材質/尺寸超出生產範圍",        "客戶",  "—",        "✅ POC",  ""),
    ("FL008", "客戶明確要求真人",     "客戶輸入要求轉真人",          "客戶",  "—",        "✅ POC",  "立即轉接，不需觸發次數"),
    ("FL009", "英文詢價",             "偵測英文輸入",                 "客戶",  "≤ 15 輪",  "✅ POC",  "以英文回應"),
    ("FL010", "LINE 詢價",            "LINE 官方帳號 Webhook",        "客戶",  "—",        "Phase 2", "需開通 LINE Business"),
]


# ══════════════════════════════════════════════════════════════════
# Sheet 3 — 欄位定義
# ══════════════════════════════════════════════════════════════════
FIELDS = [
    # (類別, 編號, 欄位名稱, 收集目的, AI 引導話術, 資料型態, 加密保護, 備註)
    ("A 類（必問）", "A01", "品名",       "識別訂單 / 推斷容器類型",   "「請問這張貼紙是用在什麼產品上呢？」",                                            "文字",          "否",  "同時推斷容器類型用於模擬圖"),
    ("A 類（必問）", "A02", "尺寸",       "計算用料",                  "「貼紙大概是幾公分寬、幾公分高呢？」",                                            "數值 (mm × mm)", "否",  "成捲需確認出紙方向"),
    ("A 類（必問）", "A03", "數量",       "報價基礎",                  "「這次大概需要印幾張呢？」",                                                      "整數 (PCS)",    "否",  "影響單價區間"),
    ("A 類（必問）", "A04", "材質",       "工法與用料",                "「您希望貼紙是亮面的、霧面的，還是有特殊質感（如銀色金屬感）？」",                "Enum（8 種）",  "否",  "限官網 8 種，見「材質對照表」"),
    ("A 類（必問）", "A05", "色數",       "估算印刷成本",              "AI 從設計稿解析；或問：「設計上有幾個顏色？有沒有特別調配的專色？」",            "整數",          "否",  "CMYK + 特色數量"),
    ("A 類（必問）", "A06", "成捲 / 單張", "後加工方式",               "「這批貼紙要裁成一張一張，還是捲成一卷（方便機器自動貼）？」",                   "Enum",          "否",  "成捲需定義捲向"),
    ("A 類（必問）", "A07", "是否軋型",   "是否需刀模",                "「貼紙形狀是長方形、圓形，還是需要特別裁切成產品輪廓形狀？」",                    "Boolean",       "否",  "非方形需刀模費"),
    ("A 類（必問）", "A08", "出貨日期",   "排定生產優先序",            "「您希望什麼時候可以收到貨呢？」",                                                "日期",          "否",  "影響生產排程"),
    ("A 類（必問）", "A09", "交貨方式",   "物流安排",                  "「請問要送貨到府，還是您會自行來取？」",                                          "Enum",          "否",  "含物流廠商選擇"),
    ("A 類（必問）", "A10", "收件人",     "送貨資訊",                  "「最後請留下收件人姓名、地址和聯絡電話，我們安排出貨用。」",                      "文字",          "✅ 加密", "個資，需加密儲存"),
    ("A 類（必問）", "A10", "地址",       "送貨資訊",                  "（同上）",                                                                        "文字",          "✅ 加密", "個資，需加密儲存"),
    ("A 類（必問）", "A10", "聯絡電話",   "送貨資訊",                  "（同上）",                                                                        "文字",          "✅ 加密", "個資，需加密儲存"),
    ("A 類（必問）", "A11", "發票抬頭",   "開立統一發票",              "「發票要開給哪個公司行號呢？」",                                                  "文字",          "否",  "B2B 必問"),
    ("A 類（必問）", "A12", "客戶生產單號", "大客戶內部追蹤",          "「您們有沒有自己的採購單號需要我們標注？（沒有的話也沒關係）」",                  "文字 / null",   "否",  "中小客戶可略過"),
    ("A 類（必問）", "A13", "客戶料號",   "對應客戶編碼",              "「這款貼紙在您們系統裡有沒有對應的料號？」",                                      "文字 / null",   "否",  "大客戶通常有"),
    ("A 類（必問）", "A14", "接單日期",   "訂單時間追蹤",              "系統自動填入，不問客戶",                                                          "日期 (auto)",   "否",  ""),
    ("B 類（條件）", "B01", "圓角",       "",                          "「貼紙的四個角要圓一點嗎？還是直角就好？」",                                      "文字 / null",   "否",  "觸發：客戶提到形狀或裁切"),
    ("B 類（條件）", "B02", "亮 / 霧膜",  "",                          "「表面要亮面（反光）還是霧面（低調）？」",                                        "Enum / null",   "否",  "觸發：客戶提到表面質感"),
    ("B 類（條件）", "B03", "燙金色號 / 尺寸", "",                     "「有需要燙上金色或銀色的反光效果嗎？大概在哪個位置？」",                          "文字 / null",   "否",  "觸發：提到金色／銀色／反光"),
    ("B 類（條件）", "B04", "打凸",       "",                          "「有需要讓某個部分凸起來，形成立體觸感嗎？」",                                    "Boolean / null", "否", "觸發：提到立體感"),
    ("B 類（條件）", "B05", "成樣方式",   "",                          "「這批貼紙出貨時要怎麼包裝或分裝呢？」",                                          "文字 / null",   "否",  "觸發：出貨形式不標準"),
    ("B 類（條件）", "B06", "模擬圖打樣", "",                          "「要不要先看看貼在產品上的模擬效果圖呢？」",                                      "Boolean",       "否",  "觸發：A 類欄位收集 50%+"),
    ("B 類（條件）", "B07", "備註",       "",                          "「還有其他需要特別說明的嗎？」",                                                  "文字 / null",   "否",  "觸發：對話中有特殊需求"),
]


# ══════════════════════════════════════════════════════════════════
# Sheet 4 — 材質對照表
# ══════════════════════════════════════════════════════════════════
MATERIALS = [
    # (材質名稱, 特性描述, 適合場景, AI 白話說法, 防水, 抗刮, 備註)
    ("銅版紙",     "光滑平面，CP值高",             "一般標籤、預算考量",         "「最基本款，價格實惠，表面光滑」",               "否",  "否",  ""),
    ("亮面珠光",   "色彩表現佳，防水、撕不破",     "保養品、飲料、禮品",         "「光澤感強，色彩鮮豔，防水耐用」",               "✅", "✅", ""),
    ("霧面珠光紙", "質感佳，消光效果，防水",        "精品、美妝、高端酒標",       "「低調質感，霧面不反光，同樣防水」",             "✅", "✅", ""),
    ("亮銀龍貼紙", "防水、抗刮，鏡面表面",          "電子產品、金屬質感需求",     "「鏡面銀色，像不鏽鋼一樣的質感」",              "✅", "✅", ""),
    ("反銀龍貼紙", "防水、抗刮，霧面銀",            "工業標籤、耐久性需求",       "「銀色但較低調，同樣防水抗刮」",                "✅", "✅", ""),
    ("模造紙",     "非塗佈紙，吸墨強，無光澤",      "手寫標籤、有機風格",         "「無光澤、自然質感，墨水吸收好，可書寫」",       "否",  "否",  ""),
    ("透明標籤",   "防水、抗刮、撕不破，透明",      "瓶罐透視設計",               "「貼上去幾乎看不到底材，像直接印在容器上」",     "✅", "✅", ""),
    ("日本和紙",   "質感顯著，質地特殊，紋路多樣",  "高端禮品、清酒標、文創",     "「有獨特紙感和紋路，非常有質感，適合精品」",     "否",  "否",  ""),
]


# ══════════════════════════════════════════════════════════════════
# Sheet 5 — 角色清單
# ══════════════════════════════════════════════════════════════════
ROLES = [
    ("客戶",        "詢價、下訂、補印的終端使用者",                     "只能使用 Chatbot 介面，無後台存取權限"),
    ("業務",        "接收草稿通知、確認訂單、處理人工轉接",              "使用管理後台，可查看草稿列表與轉接紀錄"),
    ("業務主管",    "監督訂單流程、查看統計報表、管理業務帳號",           "使用管理後台，較業務有更高查閱權限"),
    ("系統管理員",  "系統配置、使用者帳號管理、Odin 工作流維護",         "最高操作權限，負責 Odin 工作流與後台設定"),
    ("系統",        "Odin 工作流、AP Server 等自動化元件的操作主體",     "非人員角色，代表後端自動化作業"),
]


# ══════════════════════════════════════════════════════════════════
# Sheet 6 — 訪談問題清單
# ══════════════════════════════════════════════════════════════════
QUESTIONS = [
    # (主題, 優先序, 編號, 問題, 背景說明, 確認結果)
    ("業務流程現況", "P1", "Q01", "目前散客主要透過哪個管道詢價？（官網表單 / Email / LINE / 電話）",         "系統要優先對接哪個管道",                                   ""),
    ("業務流程現況", "P1", "Q02", "目前一天平均收到幾筆詢價？旺季與淡季差異大嗎？",                          "評估系統承載量與成本效益",                                 ""),
    ("業務流程現況", "P1", "Q03", "從收到詢價到回覆報價，現在的標準作業時間是多久？",                        "設定 KPI 基準線",                                          ""),
    ("業務流程現況", "P1", "Q04", "詢價到成立訂單的轉換率大約是多少？流失主因是什麼？",                      "量化導入效益",                                             ""),
    ("業務流程現況", "P1", "Q05", "目前詢價到下訂，通常要來回幾次才能確認完規格？",                          "設定「≤ 15 輪」目標的合理性",                              ""),
    ("業務流程現況", "P2", "Q06", "是否有跨國或英文詢價需求？佔比大約多少？",                               "評估多語言功能的優先序",                                   ""),
    ("範疇與優先序", "P0", "Q07", "第一個上線的管道優先選哪一個？（官網嵌入 chatbot vs. LINE 官方帳號）",    "影響 Phase 1 開發範疇",                                    ""),
    ("範疇與優先序", "P1", "Q08", "圖檔上傳與 AI 分析功能是否列入本次 POC？",                              "技術複雜度較高，但是差異化亮點",                           ""),
    ("範疇與優先序", "P1", "Q09", "AI 貼標模擬圖生成是否為本次驗收的必要功能，還是 nice-to-have？",          "釐清驗收標準",                                             ""),
    ("範疇與優先序", "P1", "Q10", "業務後台（草稿確認介面）是否列入本次 POC？還是以 Email 通知 + 人工查看 JSON 即可？", "影響開發工作量",                              ""),
    ("範疇與優先序", "P2", "Q11", "管理後台需要帳號與權限管理功能嗎？（多位業務各自登入）",                  "影響後台開發範疇",                                         ""),
    ("範疇與優先序", "P2", "Q12", "POC 之後優先推進 LINE 整合還是自動報價？",                              "了解客戶規劃，協助排序",                                   ""),
    ("材質與規格",   "P1", "Q13", "官網 8 種材質是否為完整目錄？未來是否有新增材質的計畫？",                 "AI 只能引導官網材質，需確認清單穩定性",                    ""),
    ("材質與規格",   "P1", "Q14", "客戶詢問目錄外材質，現在的處理方式是什麼？",                             "設計「超規格轉接」的觸發條件",                             ""),
    ("材質與規格",   "P1", "Q15", "貼紙尺寸有沒有最大或最小限制？超出範圍如何處理？",                       "設計超規格觸發條件",                                       ""),
    ("材質與規格",   "P2", "Q16", "成捲捲向（正捲右出 / 左出 / 下出）客戶通常懂嗎？需要 AI 提供圖示說明嗎？", "設計 B 類欄位的引導話術",                               ""),
    ("舊客識別",     "P1", "Q17", "舊客識別以 Email 還是電話為主要比對依據？兩者都可以嗎？",                 "設計比對邏輯",                                             ""),
    ("舊客識別",     "P0", "Q18", "歷史訂單資料目前存放在哪裡？格式如何？可以提供嗎？",                      "評估舊客資料匯入的工程量",                                 ""),
    ("個資與安全",   "P0", "Q19", "是否同意將客戶個資（姓名、地址、電話）加密儲存於 Asgard AI 的 DB 中？",   "確認個資授權範圍，POC 必要問題",                           ""),
    ("個資與安全",   "P2", "Q20", "對話紀錄保留多久？有沒有定期清除的需求？",                               "資料保存政策",                                             ""),
    ("通知與轉接",   "P1", "Q21", "業務收到轉接通知的偏好管道？Email 還是 LINE 群組？還是兩者都要？",         "設計通知機制",                                             ""),
    ("通知與轉接",   "P1", "Q22", "情緒偵測觸發轉接的閾值：「連續 2 次明確不滿語句」是否合理？",             "避免假陽性或客戶流失",                                     ""),
    ("通知與轉接",   "P2", "Q23", "高金額訂單轉人工的金額門檻是多少？",                                     "需杰隆提供定價邏輯才能定義",                               ""),
    ("通知與轉接",   "P2", "Q24", "業務收到通知後，預期多快要回應客戶？有沒有 SLA 要求？",                   "設定通知 Email 的時效提示",                                ""),
    ("Excel 銜接",   "P0", "Q25", "可以提供最新版訂單追蹤 Excel？確認 AI 輸出欄位與 Excel 格式一致。",       "確認訂單 JSON Schema 與 Excel 格式對齊",                   ""),
    ("Excel 銜接",   "P1", "Q26", "POC 只收集 A 類 14 項；其餘 B 類 + C 類內部欄位由誰補填？",              "確認 POC 後的業務作業流程",                                ""),
    ("Excel 銜接",   "P1", "Q27", "AI 草稿要寫回現有 Excel，還是新開一個系統管理訂單？",                     "影響資料層設計與整合工程量",                               ""),
    ("維運管理",     "P1", "Q28", "知識庫（FAQ、材質說明）上線後，由誰負責日常維護與更新？",                  "確認維運責任歸屬",                                         ""),
    ("維運管理",     "P2", "Q29", "若 AI 給出錯誤的材質建議，設想的糾正機制是什麼？",                        "建立內容審查流程",                                         ""),
    ("維運管理",     "P2", "Q30", "系統上線後，業務人員需要哪些教育訓練？",                                  "規劃上線準備事項",                                         ""),
    ("時程與決策",   "P1", "Q31", "POC 驗收後，預計多快要上線正式版本？有無對外公告的時間壓力？",             "規劃 Phase 2 時程",                                        ""),
    ("時程與決策",   "P0", "Q32", "本次 POC 的主要決策窗口是誰？",                                          "確認簽核流程，避免後期卡關",                               ""),
    ("時程與決策",   "P1", "Q33", "杰隆官網是否允許嵌入第三方腳本 / iframe？技術上有沒有限制？",              "評估網頁嵌入可行性",                                       ""),
    ("時程與決策",   "P2", "Q34", "是否有競品或參考案例希望我們對標？",                                      "了解期望值與品質標準",                                     ""),
]


# ══════════════════════════════════════════════════════════════════
# 建構 Excel
# ══════════════════════════════════════════════════════════════════
def build_excel(output_path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1：功能地圖 ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "功能地圖"
    cols1 = [
        ("ID", 6), ("平台", 22), ("功能", 20), ("Epic", 16),
        ("子功能", 22), ("功能說明", 52), ("可使用角色", 18), ("狀態", 10), ("備註", 24),
    ]
    header_row(ws1, cols1)
    ws1.row_dimensions[1].height = 22

    prev_platform = None
    for r_idx, row in enumerate(FUNC_MAP, 2):
        platform = row[1]
        if platform != prev_platform:
            for c in range(1, 10):
                cell = ws1.cell(row=r_idx, column=c, value=row[c - 1])
                cell.fill = GRP_FILL
                cell.font = GRP_FONT
                cell.alignment = WRAP
                cell.border = BORDER
            prev_platform = platform
        else:
            fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
            for c_idx, val in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == 8:  # 狀態欄
                    if val == "✅ POC":
                        cell.fill = STATUS_OK
                    elif val == "Phase 2":
                        cell.fill = STATUS_P2
                    else:
                        cell.fill = STATUS_PND
                    cell.alignment = CENTER
                else:
                    cell.fill = fill
                    cell.alignment = WRAP
                cell.font = BODY_FONT
                cell.border = BORDER
        ws1.row_dimensions[r_idx].height = 40

    freeze_and_filter(ws1)
    ws1.sheet_view.showGridLines = False

    # ── Sheet 2：對話流程清單 ──────────────────────────────────
    ws2 = wb.create_sheet("對話流程清單")
    cols2 = [
        ("流程編號", 10), ("流程名稱", 28), ("觸發條件", 32),
        ("使用角色", 12), ("對話步驟上限", 14), ("狀態", 10), ("備註", 26),
    ]
    header_row(ws2, cols2)
    for r_idx, row in enumerate(FLOW_LIST, 2):
        fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 6:
                cell.fill = STATUS_OK if val == "✅ POC" else STATUS_P2
                cell.alignment = CENTER
            else:
                cell.fill = fill
                cell.alignment = WRAP
            cell.font = BODY_FONT
            cell.border = BORDER
        ws2.row_dimensions[r_idx].height = 30
    freeze_and_filter(ws2)
    ws2.sheet_view.showGridLines = False

    # ── Sheet 3：欄位定義 ──────────────────────────────────────
    ws3 = wb.create_sheet("欄位定義")
    cols3 = [
        ("類別", 12), ("編號", 7), ("欄位名稱", 16), ("收集目的", 20),
        ("AI 引導話術", 44), ("資料型態", 18), ("個資加密", 10), ("備註", 26),
    ]
    header_row(ws3, cols3)
    prev_cat = None
    for r_idx, row in enumerate(FIELDS, 2):
        cat = row[0]
        if cat != prev_cat:
            for c in range(1, 9):
                cell = ws3.cell(row=r_idx, column=c, value=row[c - 1])
                cell.fill = GRP_FILL
                cell.font = GRP_FONT
                cell.alignment = WRAP
                cell.border = BORDER
            prev_cat = cat
        else:
            fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
            for c_idx, val in enumerate(row, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == 7 and val == "✅ 加密":
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    cell.alignment = CENTER
                else:
                    cell.fill = fill
                    cell.alignment = WRAP
                cell.font = BODY_FONT
                cell.border = BORDER
        ws3.row_dimensions[r_idx].height = 44
    freeze_and_filter(ws3)
    ws3.sheet_view.showGridLines = False

    # ── Sheet 4：材質對照表 ──────────────────────────────────
    ws4 = wb.create_sheet("材質對照表")
    cols4 = [
        ("材質名稱", 16), ("特性描述", 26), ("適合場景", 26),
        ("AI 白話說法", 34), ("防水", 8), ("抗刮", 8), ("備註", 20),
    ]
    header_row(ws4, cols4)
    for r_idx, row in enumerate(MATERIALS, 2):
        fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
        for c_idx, val in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.alignment = WRAP if c_idx != 5 else CENTER
            cell.font = BODY_FONT
            cell.border = BORDER
        ws4.row_dimensions[r_idx].height = 36
    freeze_and_filter(ws4)
    ws4.sheet_view.showGridLines = False

    # ── Sheet 5：角色清單 ──────────────────────────────────────
    ws5 = wb.create_sheet("角色清單")
    cols5 = [("系統角色名稱", 18), ("描述 / 說明", 44), ("備註", 44)]
    header_row(ws5, cols5)
    for r_idx, row in enumerate(ROLES, 2):
        fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
        for c_idx, val in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.alignment = WRAP
            cell.font = BODY_FONT
            cell.border = BORDER
        ws5.row_dimensions[r_idx].height = 36
    freeze_and_filter(ws5)
    ws5.sheet_view.showGridLines = False

    # ── Sheet 6：訪談問題清單 ──────────────────────────────────
    ws6 = wb.create_sheet("訪談問題清單")
    cols6 = [
        ("主題", 16), ("優先序", 8), ("編號", 7),
        ("問題", 52), ("背景說明", 36), ("確認結果（訪談後填入）", 36),
    ]
    header_row(ws6, cols6)
    PRIORITY_FILL = {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL}
    prev_topic = None
    for r_idx, row in enumerate(QUESTIONS, 2):
        topic = row[0]
        prio  = row[1]
        if topic != prev_topic:
            for c in range(1, 7):
                cell = ws6.cell(row=r_idx, column=c, value=row[c - 1])
                cell.fill = GRP_FILL
                cell.font = GRP_FONT
                cell.alignment = WRAP
                cell.border = BORDER
            prev_topic = topic
        else:
            fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
            for c_idx, val in enumerate(row, 1):
                cell = ws6.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == 2:
                    cell.fill = PRIORITY_FILL.get(prio, fill)
                    cell.alignment = CENTER
                    cell.font = Font(bold=True, name="Calibri", size=10)
                else:
                    cell.fill = fill
                    cell.alignment = WRAP
                    cell.font = BODY_FONT
                cell.border = BORDER
        ws6.row_dimensions[r_idx].height = 44
    freeze_and_filter(ws6)
    ws6.sheet_view.showGridLines = False

    wb.save(output_path)
    print(f"✅ 已產生：{output_path}")


if __name__ == "__main__":
    out = Path(__file__).parent.parent / "杰隆印刷_AI接單系統_FunctionMap_Questions_v0.1.xlsx"
    build_excel(str(out))
